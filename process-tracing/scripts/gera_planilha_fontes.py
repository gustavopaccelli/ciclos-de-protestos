"""gera_planilha_fontes.py — Gera fontes-de-dados.xlsx a partir do CSV canônico.

xlsx é binário: o git não versiona diferença legível dele. Por isso a fonte de
verdade é fontes/fontes-de-dados.csv, versionável e diffável, e o .xlsx é
derivado — mesmo padrão de artigo/referencias.bib -> referencias-abnt.md.

NÃO edite o .xlsx: a próxima execução sobrescreve.

Uso:
  python gera_planilha_fontes.py
  python gera_planilha_fontes.py --check   # sai 1 se o xlsx estiver desatualizado
"""

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
CSV = BASE / "fontes" / "fontes-de-dados.csv"
XLSX = BASE / "fontes" / "fontes-de-dados.xlsx"

# As quatro colunas pedidas vêm primeiro; as demais dão utilidade operacional.
ORDEM = ["tipo_de_dado", "titulo", "citacao_abnt", "link", "orgao_responsavel",
         "nivel_fonte", "cobertura_temporal", "forma_de_acesso", "ciclos_atendidos",
         "data_consulta", "status_verificacao", "observacoes"]
ROTULOS = {
    "tipo_de_dado": "Tipo de dado", "titulo": "Título", "citacao_abnt": "Citação (ABNT)",
    "link": "Link", "orgao_responsavel": "Órgão responsável", "nivel_fonte": "Nível",
    "cobertura_temporal": "Cobertura temporal", "forma_de_acesso": "Forma de acesso",
    "ciclos_atendidos": "Ciclos", "data_consulta": "Consulta",
    "status_verificacao": "Status", "observacoes": "Observações",
}
LARGURA = {"tipo_de_dado": 26, "titulo": 46, "citacao_abnt": 82, "link": 54,
           "orgao_responsavel": 26, "nivel_fonte": 7, "cobertura_temporal": 24,
           "forma_de_acesso": 24, "ciclos_atendidos": 14, "data_consulta": 12,
           "status_verificacao": 24, "observacoes": 70}

TINTA = {"verificado_fonte_oficial": "FFE8F3ED", "pendente": "FFFDF6E3",
         "divergencia_nao_resolvida": "FFFBEAEC", "corroborado_imprensa": "FFEDF1F8"}
FINA = Side(style="thin", color="FFD9DDE3")
BORDA = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)


def le() -> list[dict]:
    with CSV.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def monta(linhas: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fontes de dados"

    ws["A1"] = ("Fontes de dados do levantamento por process tracing — "
                "projeto Ciclos de Protesto no Brasil")
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("ARQUIVO DERIVADO — gerado de fontes/fontes-de-dados.csv por "
                "scripts/gera_planilha_fontes.py. Não edite aqui: a próxima geração "
                "sobrescreve. Nível 1 = ato oficial primário · 2 = registro "
                "institucional · 3 = imprensa (só corroboração) · 4 = literatura. "
                "Ver fontes/hierarquia-fontes.md.")
    ws["A2"].font = Font(italic=True, size=9, color="FF5B6270")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ORDEM))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ORDEM))
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    cab = 4
    for j, campo in enumerate(ORDEM, start=1):
        c = ws.cell(row=cab, column=j, value=ROTULOS[campo])
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF2E4A78")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = LARGURA[campo]

    for i, linha in enumerate(linhas, start=cab + 1):
        tinta = TINTA.get(linha.get("status_verificacao", ""), "FFFFFFFF")
        for j, campo in enumerate(ORDEM, start=1):
            valor = linha.get(campo, "") or ""
            c = ws.cell(row=i, column=j, value=valor)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDA
            c.fill = PatternFill("solid", fgColor=tinta)
            if campo == "link" and valor.startswith("http"):
                c.hyperlink = valor
                c.font = Font(color="FF2E4A78", underline="single")
            elif campo == "nivel_fonte":
                c.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = ws.cell(row=cab + 1, column=1)
    ws.auto_filter.ref = (f"A{cab}:{get_column_letter(len(ORDEM))}{cab + len(linhas)}")
    return wb


def main(check: bool) -> None:
    linhas = le()
    if check:
        if not XLSX.exists():
            print("xlsx ausente — rode sem --check para gerar.")
            sys.exit(1)
        print(f"{len(linhas)} fontes no CSV; xlsx presente em {XLSX.name}.")
        print("Observação: xlsx é binário e não se compara byte a byte de forma estável; "
              "regenere sempre que o CSV mudar.")
        return
    monta(linhas).save(XLSX)
    print(f"{len(linhas)} fontes → {XLSX}")
    por_nivel = {}
    for l in linhas:
        por_nivel[l["nivel_fonte"]] = por_nivel.get(l["nivel_fonte"], 0) + 1
    print("  por nível: " + " · ".join(f"n{k}={v}" for k, v in sorted(por_nivel.items())))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true")
    main(ap.parse_args().check)
